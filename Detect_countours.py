import sys
from typing import List, Any

import numpy as np
import cv2 as cv
from Defenation_for_Countours import pixel_cont, create_max, create_min, rotata
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
import os

List_coort = []

def coo(col, im):
    fn = im # путь к файлу с картинкой
    img = cv.imread(fn)
    a = pixel_cont(im)
    for i in range(col):

        bgr_min = np.array(create_min(a[i]), np.uint8)
        bgr_max = np.array(create_max(a[i]), np.uint8)

        thresh = cv.inRange(img, bgr_min, bgr_max ) # применяем цветовой фильтр
            # ищем контуры и складируем их в переменную contours
        contours, hierarchy = cv.findContours(thresh.copy(), cv.RETR_TREE, cv.CHAIN_APPROX_SIMPLE)
            
            
        List_coort.append(list(contours))
        print(List_coort)
        cv.drawContours( img, contours, -1, (0,255,0), 2, cv.LINE_AA, hierarchy, 1 )


    return List_coort
            

def coort_print(col, im):
    global List_coort
    fn = im # путь к файлу с картинкой
    img = cv.imread(fn)
    a = pixel_cont(im)

    for i in range(col):

        bgr_min = np.array(create_min(a[i]), np.uint8)
        bgr_max = np.array(create_max(a[i]), np.uint8)

        thresh = cv.inRange(img, bgr_min, bgr_max ) # применяем цветовой фильтр
            # ищем контуры и складируем их в переменную contours
        contours, hierarchy = cv.findContours( thresh.copy(), cv.RETR_TREE, cv.CHAIN_APPROX_SIMPLE)
            
        List_coort.append(list(contours))
        print(List_coort)
        cv.drawContours(img, contours, -1, (0,255,0), 2, cv.LINE_AA, hierarchy, 1)
    
    
    cv.imshow("girl", img)
    cv.waitKey(0)
    
    return List_coort

def raspk_coort(coort):
    # Функция для распаковки координат из numpy архива
    final_res = []
    final_res1 = []
    res = []
    for i in coort:       
        for j in i:
            res.append(j.tolist())
    for i in res:

        final_res.append(i[-2])
    for i in final_res:
        for j in i:
            
            final_res1.append(j)


    return sorted(final_res1)



def cont_to_ex(name, coort):
    # Конвертация точек в график excel
    wb = Workbook()
    wb.create_sheet(title=name, index=0)

    shhet = wb[name]

    nw = []

    for i in coort:
        nw.append(i[1])
    nw.reverse()

    # Записываем значения точек в ячейки в обратном порядке 
    for i in range(len(nw)):
        cell = shhet.cell(row=i + 1, column=1)

        cell.value = nw[i]
       
        print(cell.value)

    chart = BarChart()
    chart.title = name + " Convert"

    # Создаём пропорциональный изображению график
    data = Reference(shhet, min_col=1, min_row=1, max_col=1, max_row=len(coort))


    chart.add_data(data)

    shhet.add_chart(chart, 'C2')

    wb.save(name + '.xlsx')

    List_coort.clear()
    return name + '.xlsx'



# # Функция для ковертации точек
# coort_print(len(pixel_cont(rotata(r"D:\project\images\6.png"))), rotata(r"D:\project\images\6.png"))



# cont_to_ex("jlewhfv111we", raspk_coort(List_coort))


# os.remove(rotata("2.png"))