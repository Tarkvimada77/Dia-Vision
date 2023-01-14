import cv2
from openpyxl import Workbook
from openpyxl.chart import Reference
from openpyxl.chart import PieChart


def radial_convert(puth, name):
    img = cv2.imread(puth)

    img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    ret, thresh = cv2.threshold(img_gray, 170, 255, cv2.THRESH_BINARY)

    contours, hierarchy = cv2.findContours(image=thresh, mode=cv2.RETR_TREE, method=cv2.CHAIN_APPROX_NONE)

    image_copy = img.copy()
    cv2.drawContours(image=image_copy, contours=contours, contourIdx=-1, color=(0, 255, 0), thickness=2,
                     lineType=cv2.LINE_AA)

    coort = list(contours)
    coort.pop(0)

    S_list = []
    for i in coort:
        S_list.append(int(cv2.contourArea(i, oriented=False)))
    S_list.sort()

    # Конвертация точек в график excel
    wb = Workbook()
    wb.create_sheet(title=name, index=0)

    shhet = wb[name]

    for i in range(len(S_list)):
        cell = shhet.cell(row=i + 1, column=1)

        cell.value = S_list[i]

    chart = PieChart()
    chart.title = name + " Convert"

    # Создаём пропорциональный изображению график
    data = Reference(shhet, min_col=1, min_row=1, max_col=1, max_row=len(S_list))

    chart.add_data(data)

    shhet.add_chart(chart, 'C2')

    wb.save(name + '.xlsx')

    coort.clear()
    S_list.clear()
    return name + '.xlsx'


# radial_convert(r"D:\Dia-Vision\Screenshot_1.png", "Pop It")

